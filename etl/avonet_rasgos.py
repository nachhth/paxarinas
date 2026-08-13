"""Fonte AVONET: tamaño e hábitat de cada especie, para poder buscar sen o nome.

O uso real da app é o contrario do que resolve un catálogo alfabético: alguén
ve un paxaro, non sabe como se chama, e quere chegar a el. Para iso fan falta
rasgos, e os rasgos non están en GBIF.

AVONET (Tobias et al. 2022, Ecology Letters) mide 11.009 especies: masa, á,
tarso, bico, e ademais hábitat e nicho trófico. Licenza CC BY 4.0.

Do que ofrece só se colle o que lle serve a alguén que mira un paxaro: canto
mide, canto pesa, onde vive e de que come. As medidas de bico e tarso quedan
fóra: son excelentes para un traballo de morfoloxía e inútiles no campo.

Uso:
    python etl/avonet_rasgos.py

Saída:
    etl/out/avonet_rasgos.json
"""

from __future__ import annotations

import json
import zipfile
from pathlib import Path

from common import (CACHE_DIR, OUT_DIR, descarga_ficheiro, escribe_json,
                    get_json, log)

FIGSHARE = "https://api.figshare.com/v2/articles/16586228/files"
FICHEIRO = "AVONET Supplementary dataset 1.xlsx"

# As tres follas traen os mesmos datos baixo tres taxonomías distintas. Úsanse
# as tres porque os nosos nomes veñen do backbone de GBIF, que non coincide con
# ningunha: o que non casa nunha adoita casar noutra.
FOLLAS = {
    "AVONET1_BirdLife": ("Species1", "Habitat", "Trophic.Niche", "Mass",
                         "Wing.Length", "Tarsus.Length", "Tail.Length"),
    "AVONET2_eBird": ("Species2", "Habitat", "Trophic.Niche", "Mass",
                      "Wing.Length", "Tarsus.Length", "Tail.Length"),
    "AVONET3_BirdTree": ("Species3", "Habitat", "Trophic.Niche", "Mass",
                         "Wing.Length", "Tarsus.Length", "Tail.Length"),
}

# Referencias que calquera recoñece, para non dar gramos a secas. A xente non
# estima pesos, compara: "coma un merlo" di moito máis que "90 gramos".
REFERENCIAS = [
    (12, "moi pequena", "coma un ferreiriño"),
    (30, "pequena", "coma un pardal"),
    (100, "mediana", "coma un merlo"),
    (350, "grande", "coma unha pomba"),
    (1200, "moi grande", "coma un corvo ou unha gaivota"),
    (float("inf"), "enorme", "coma un ganso ou unha aguia"),
]

HABITATS = {
    "Forest": "bosque",
    "Woodland": "fraga",
    "Shrubland": "matogueira",
    "Grassland": "pasteiro",
    "Wetland": "zona húmida",
    "Riverine": "ribeira",
    "Coastal": "costa",
    "Marine": "mar aberto",
    "Rock": "rochedo",
    "Desert": "deserto",
    "Human Modified": "zonas humanizadas",
}

NICHOS = {
    "Invertivore": "insectos e invertebrados",
    "Vertivore": "outros vertebrados",
    "Aquatic predator": "peixes e presas acuáticas",
    "Frugivore": "froita",
    "Granivore": "sementes",
    "Nectarivore": "néctar",
    "Herbivore aquatic": "plantas acuáticas",
    "Herbivore terrestrial": "plantas",
    "Omnivore": "de todo",
    "Scavenger": "carroña",
}


def clase_de_tamano(masa: float | None) -> tuple[str | None, str | None]:
    if masa is None:
        return None, None
    for limite, clase, comparanza in REFERENCIAS:
        if masa <= limite:
            return clase, comparanza
    return None, None


def descarga_avonet() -> Path:
    destino = CACHE_DIR / "avonet.xlsx"
    if destino.exists() and destino.stat().st_size > 0:
        return destino

    log("Descargando AVONET de figshare (20 MB)...")
    ficheiros = get_json(FIGSHARE)
    url = next((f["download_url"] for f in ficheiros if f["name"] == FICHEIRO), None)
    if not url:
        raise SystemExit(f"Non se atopa '{FICHEIRO}' en figshare.")

    descarga_ficheiro(url, destino)
    return destino


def le_follas(ruta: Path) -> dict[str, dict]:
    """nome científico -> rasgos. As tres taxonomías van ao mesmo dicionario."""
    import openpyxl

    if zipfile.is_zipfile(ruta) is False:
        raise SystemExit(f"{ruta} non parece un xlsx válido. Bórrao e reinténtao.")

    libro = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    rasgos: dict[str, dict] = {}

    for folla, columnas in FOLLAS.items():
        if folla not in libro.sheetnames:
            log(f"  aviso: falta a folla {folla}")
            continue

        ws = libro[folla]
        filas = ws.iter_rows(values_only=True)
        cabeceira = [str(c).strip() if c else "" for c in next(filas)]
        indice = {nome: cabeceira.index(nome) for nome in columnas
                  if nome in cabeceira}

        if len(indice) < len(columnas):
            log(f"  aviso: en {folla} faltan columnas: "
                f"{set(columnas) - set(indice)}")

        col_nome = columnas[0]
        n = 0
        for fila in filas:
            nome = fila[indice[col_nome]] if col_nome in indice else None
            if not nome:
                continue
            nome = str(nome).strip()
            # A primeira taxonomía que dea un nome mándao: non se sobrescribe.
            if nome in rasgos:
                continue

            def valor(clave):
                return fila[indice[clave]] if clave in indice else None

            def numero(clave):
                v = valor(clave)
                try:
                    return round(float(v), 1)
                except (TypeError, ValueError):
                    return None

            rasgos[nome] = {
                "masa": numero("Mass"),
                "ala": numero("Wing.Length"),
                "tarso": numero("Tarsus.Length"),
                "cola": numero("Tail.Length"),
                "habitat": valor("Habitat"),
                "nicho": valor("Trophic.Niche"),
            }
            n += 1

        log(f"  {folla}: {n} especies novas (total {len(rasgos)})")

    libro.close()
    return rasgos


def main() -> None:
    fonte = OUT_DIR / "gbif_especies.json"
    if not fonte.exists():
        raise SystemExit(f"Falta {fonte}. Executa antes: python etl/gbif_especies.py")

    ruta = descarga_avonet()
    log("Lendo as tres taxonomías de AVONET...")
    todos = le_follas(ruta)
    log(f"AVONET: {len(todos)} nomes científicos distintos.\n")

    especies = json.loads(fonte.read_text(encoding="utf-8"))["especies"]

    catalogo: dict[str, dict] = {}
    sen_rasgos: list[str] = []

    for e in especies:
        sci = e["nomeCientifico"]
        r = todos.get(sci)
        if not r:
            sen_rasgos.append(sci)
            continue

        clase, comparanza = clase_de_tamano(r["masa"])
        catalogo[sci] = {
            "masa": r["masa"],
            "ala": r["ala"],
            "tamano": clase,
            "comparanza": comparanza,
            "habitat": HABITATS.get(r["habitat"] or "", None),
            "habitatOrixe": r["habitat"],
            "come": NICHOS.get(r["nicho"] or "", None),
            "nichoOrixe": r["nicho"],
        }

    destino = escribe_json("avonet_rasgos.json", {
        "fonte": "AVONET (Tobias et al. 2022, Ecology Letters)",
        "licenza": "CC BY 4.0",
        "url": "https://doi.org/10.1111/ele.13898",
        "total": len(catalogo),
        "rasgos": catalogo,
    })

    por_tamano: dict[str, int] = {}
    por_habitat: dict[str, int] = {}
    for r in catalogo.values():
        if r["tamano"]:
            por_tamano[r["tamano"]] = por_tamano.get(r["tamano"], 0) + 1
        if r["habitat"]:
            por_habitat[r["habitat"]] = por_habitat.get(r["habitat"], 0) + 1

    log(f"Especies con rasgos: {len(catalogo)} de {len(especies)}")
    log(f"Sen rasgos:          {len(sen_rasgos)}")
    log("\nPor tamaño:")
    for clase, _, _ in [(c, 0, 0) for _, c, _ in REFERENCIAS]:
        if clase in por_tamano:
            log(f"  {clase:12} {por_tamano[clase]:4}")
    log("\nPor hábitat:")
    for h, n in sorted(por_habitat.items(), key=lambda x: -x[1]):
        log(f"  {h:20} {n:4}")

    if sen_rasgos:
        log(f"\nExemplos sen rasgos: {', '.join(sen_rasgos[:8])}")
    log(f"\nEscrito en {destino}")


if __name__ == "__main__":
    main()
